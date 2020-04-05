import openpyxl as xl; 
df = pd.read_excel (r'D:\BothOfUs\pytoh\Book1.xlsx')
print (df)
filename ="D:\BothOfUs\pytoh\Book1.xlsx"
wb1 = xl.load_workbook(filename) 
ws1 = wb1.worksheets[0] 
mr = ws1.max_row 
mc = ws1.max_column 
for i in range (1, mr + 1): 
    for j in range (1, mc + 1):  
        c = ws1.cell(row = i, column = j)        
message=input()
poolurl="https://www.google.com"
def bot(message):
    if "/pool" in message:
            message=message.split(" ")
            del message[0]
            poolname=message
            text="Please! Give more details of the pool"
            print(poolname,text)
            message=input()
            pooldes=message
            print("Pool",poolname,"has been created - ",pooldes)
            print("Your pool is almost ready: Go to your specific pool page on, ",poolurl,"to add a catalogue and see the pool order recap.")
